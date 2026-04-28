import streamlit as st
import requests
import os
import pandas as pd
import cloudscraper
import time
import re
import glob
import hashlib
import pickle
import multiprocessing
from concurrent.futures import ThreadPoolExecutor, as_completed
import fitz  # PyMuPDF
import json
from io import BytesIO, StringIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

# Try to import the fulltext article downloader
try:
    from fulltext_article_downloader import download_article
    FULLTEXT_AVAILABLE = True
except ImportError:
    FULLTEXT_AVAILABLE = False
    st.warning("fulltext-article-downloader not installed. Run: pip install git+https://github.com/computron/fulltext-article-downloader.git")

# ---------- THEME MANAGEMENT (system only, no switcher) ----------
def apply_system_theme():
    """Apply system theme (auto follows OS/browser preference)."""
    bg = "var(--st-color-background)"
    card_bg = "var(--st-color-surface)"
    text = "var(--st-color-text)"
    border = "var(--st-color-border)"
    input_bg = "var(--st-color-surface)"
    button_bg = "var(--st-color-secondary)"
    button_hover = "var(--st-color-secondary-hover)"
    metric_gradient = "linear-gradient(135deg, #667eea, #764ba2)"

    st.markdown(f"""
    <style>
        .stApp {{ background-color: {bg}; }}
        .main .block-container {{
            background-color: {bg};
            padding-top: 2rem;
            padding-bottom: 2rem;
        }}
        .main-header {{
            font-size: 2.5rem;
            font-weight: 600;
            color: #f0f0f0;
            background: none;
            margin-top: 2rem;
            margin-bottom: 0.2rem;
        }}
        .sub-header {{
            font-size: 1.2rem;
            color: var(--st-color-text-secondary);
            margin-top: -0.5rem;
            margin-bottom: 1.5rem;
        }}
        .card {{
            background-color: {card_bg};
            border-radius: 20px;
            padding: 1.5rem;
            box-shadow: 0 4px 12px rgba(0,0,0,0.05);
            margin-bottom: 1rem;
            border: 1px solid {border};
            transition: transform 0.2s;
        }}
        .card:hover {{
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(0,0,0,0.1);
        }}
        .metric-row {{
            display: flex;
            gap: 1rem;
            margin-bottom: 2rem;
            margin-top: 1rem;
        }}
        .metric-card {{
            background: {metric_gradient};
            border-radius: 20px;
            padding: 1rem 0.5rem;
            color: white;
            text-align: center;
            flex: 1;
            box-shadow: 0 4px 12px rgba(0,0,0,0.1);
        }}
        .metric-card h3 {{
            font-size: 2rem;
            margin: 0;
            font-weight: 700;
        }}
        .metric-card p {{
            margin: 0;
            font-size: 0.9rem;
            opacity: 0.9;
        }}
        .stButton > button {{
            border-radius: 30px;
            font-weight: 500;
            background-color: {button_bg};
            color: {text};
            border: 1px solid {border};
            transition: all 0.2s;
        }}
        .stButton > button:hover {{
            transform: scale(1.02);
            background-color: {button_hover};
            color: {text};
        }}
        .stTextInput > div > div > input, 
        .stTextArea > div > div > textarea, 
        .stSelectbox > div > div {{
            border-radius: 15px;
            background-color: {input_bg};
            color: {text};
            border: 1px solid {border};
        }}
        .stProgress > div > div > div > div {{ background-color: #20c997; }}
        .streamlit-expanderHeader {{
            background-color: {card_bg};
            border-radius: 15px;
            color: {text};
            border: 1px solid {border};
        }}
        .dataframe {{
            background-color: {card_bg};
            color: {text};
        }}
        .stTabs [data-baseweb="tab-list"] {{
            gap: 8px;
            background-color: {card_bg};
            border-radius: 40px;
            padding: 6px;
            margin-bottom: 0;
        }}
        .stTabs [data-baseweb="tab"] {{
            border-radius: 30px;
            padding: 0.5rem 1.2rem;
            font-weight: 500;
            background-color: transparent;
            color: {text};
        }}
        .stTabs [aria-selected="true"] {{
            background-color: {button_bg};
            color: {text};
            box-shadow: none;
        }}
        .sidebar .sidebar-content {{ background-color: {bg}; }}
        footer {{ visibility: hidden; }}
        p, li, span, label {{ color: {text} !important; }}
        hr {{ border-color: {border}; }}
        .block-container {{ padding-top: 1rem !important; }}
        .stMarkdown {{ margin-bottom: 0.5rem; }}
    </style>
    """, unsafe_allow_html=True)

# ---------- Optional dependencies ----------
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import platform
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

def create_pdf_report(results, output_spec):
    if not REPORTLAB_AVAILABLE:
        return None, "ReportLab not installed. Please run: pip install reportlab"
    valid_results = [r for r in results if r.get("data")]
    if not valid_results:
        return None, "No valid data"
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    try:
        if platform.system() == "Windows":
            font_path = "C:/Windows/Fonts/arial.ttf"
        elif platform.system() == "Darwin":
            font_path = "/System/Library/Fonts/Supplemental/Arial Unicode.ttf"
        else:
            font_path = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
        pdfmetrics.registerFont(TTFont('UnicodeFont', font_path))
        style_normal = ParagraphStyle('CustomNormal', parent=styles['Normal'], fontName='UnicodeFont', fontSize=10)
        style_heading = ParagraphStyle('CustomHeading', parent=styles['Heading1'], fontName='UnicodeFont', fontSize=14)
    except:
        style_normal = styles['Normal']
        style_heading = styles['Heading1']
    story = []
    story.append(Paragraph("Researcher OS Analysis Report", style_heading))
    story.append(Spacer(1, 0.2*inch))
    for r in valid_results:
        story.append(Paragraph(f"<b>File:</b> {r['file']} (Pages: {r['pages']})", style_normal))
        story.append(Spacer(1, 0.1*inch))
        for k, v in r['data'].items():
            safe_k = str(k).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            safe_v = str(v).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            story.append(Paragraph(f"<b>{safe_k}:</b> {safe_v}", style_normal))
        story.append(Spacer(1, 0.2*inch))
    doc.build(story)
    pdf_data = buffer.getvalue()
    buffer.close()
    return pdf_data, None

try:
    import litellm
    from litellm import completion
    LITELLM_AVAILABLE = True
except ImportError:
    LITELLM_AVAILABLE = False

# ---------- Critical macOS Fork Safety ----------
if __name__ == '__main__':
    try:
        if multiprocessing.get_start_method(allow_none=True) != 'spawn':
            multiprocessing.set_start_method('spawn', force=True)
    except RuntimeError:
        pass
os.environ["OBJC_DISABLE_INITIALIZE_FORK_SAFETY"] = "YES"

# ---------- Setup & Session State ----------
st.set_page_config(page_title="Researcher OS v3.5 — System Theme", layout="wide", page_icon="📚")
DOWNLOAD_DIR = "./downloads"
CACHE_DIR = "./pdf_cache"
for d in [DOWNLOAD_DIR, CACHE_DIR]:
    os.makedirs(d, exist_ok=True)

# Initialize session states
if "dl_queue" not in st.session_state:
    st.session_state.dl_queue = ""
if "results_df" not in st.session_state:
    st.session_state.results_df = None
if "pdf_files" not in st.session_state:
    st.session_state.pdf_files = []
if "ai_results" not in st.session_state:
    st.session_state.ai_results = []
if "processing" not in st.session_state:
    st.session_state.processing = False
if "source_folder" not in st.session_state:
    st.session_state.source_folder = os.path.abspath(DOWNLOAD_DIR)
if "unpaywall_email" not in st.session_state:
    st.session_state.unpaywall_email = ""

# ---------- Helper Functions ----------
def clean_abstract(text):
    if not text:
        return "No abstract available."
    clean = re.sub(r'<[^>]+>', '', str(text))
    return " ".join(clean.split())

def search_openalex(query, limit=500):
    all_data = []
    per_page = min(200, limit)
    pages = (limit + per_page - 1) // per_page
    for page in range(1, pages + 1):
        url = "https://api.openalex.org/works"
        params = {"search": query, "per_page": per_page, "page": page,
                  "select": "doi,display_name,publication_year,authorships,abstract_inverted_index,host_venue"}
        try:
            r = requests.get(url, params=params, timeout=10)
            results = r.json().get('results', [])
            if not results:
                break
            for item in results:
                abstract = "No abstract available."
                index = item.get('abstract_inverted_index')
                if index:
                    words = {}
                    for word, pos_list in index.items():
                        for pos in pos_list:
                            words[pos] = word
                    abstract = " ".join([words[i] for i in sorted(words.keys())])
                authors = ", ".join([a.get('author', {}).get('display_name', '') for a in item.get('authorships', [])])
                all_data.append({
                    'Source': 'OpenAlex',
                    'Title': item.get('display_name'),
                    'Authors': authors,
                    'Journal': item.get('host_venue', {}).get('display_name', 'N/A'),
                    'Year': str(item.get('publication_year')),
                    'DOI': item.get('doi', '').replace('https://doi.org/', ''),
                    'Abstract': clean_abstract(abstract)
                })
            time.sleep(0.5)
            if len(all_data) >= limit:
                all_data = all_data[:limit]
                break
        except Exception as e:
            st.warning(f"OpenAlex page {page} error: {e}")
            break
    return all_data

def search_scopus(query, api_key, limit=500):
    if not api_key:
        return []
    all_data = []
    per_page = min(200, limit)
    start = 0
    url = "https://api.elsevier.com/content/search/scopus"
    headers = {"X-ELS-APIKey": api_key, "Accept": "application/json"}
    while start < limit:
        params = {"query": query, "count": per_page, "start": start}
        try:
            r = requests.get(url, headers=headers, params=params, timeout=10)
            data = r.json().get('search-results', {})
            entries = data.get('entry', [])
            if not entries:
                break
            for item in entries:
                all_data.append({
                    'Source': 'Scopus',
                    'Title': item.get('dc:title'),
                    'Authors': item.get('dc:creator', 'N/A'),
                    'Journal': item.get('prism:publicationName', 'N/A'),
                    'Year': item.get('prism:coverDate', '')[:4],
                    'DOI': item.get('prism:doi'),
                    'Abstract': 'Restricted by Scopus'
                })
            start += per_page
            time.sleep(0.5)
            if len(all_data) >= limit:
                all_data = all_data[:limit]
                break
        except Exception as e:
            st.warning(f"Scopus error: {e}")
            break
    return all_data

def smart_download(url, filepath, extra_headers=None):
    headers = {'User-Agent': 'Mozilla/5.0', 'Accept': 'application/pdf'}
    if extra_headers:
        headers.update(extra_headers)
    try:
        scraper = cloudscraper.create_scraper()
        response = scraper.get(url, headers=headers, timeout=30, stream=True)
        if response.status_code == 200 and 'pdf' in response.headers.get('Content-Type', '').lower():
            with open(filepath, 'wb') as f:
                f.write(response.content)
            return True, "Success"
        return False, f"Status {response.status_code}"
    except Exception as e:
        return False, str(e)

def get_download_url_and_headers(doi, scopus_key, wiley_token):
    url = f"https://link.springer.com/content/pdf/{doi}.pdf"
    headers = {}
    if "10.1016" in doi and scopus_key:
        url = f"https://api.elsevier.com/content/article/doi/{doi}"
        headers = {"X-ELS-APIKey": scopus_key, "Accept": "application/pdf"}
    elif any(x in doi for x in ["10.1111", "10.1002"]):
        if wiley_token:
            url = f"https://api.wiley.com/onlinelibrary/tdm/v1/articles/{doi}"
            headers = {"CR-Clickthrough-Client-Token": wiley_token, "Accept": "application/pdf"}
        else:
            url = f"https://onlinelibrary.wiley.com/doi/pdfdirect/{doi}"
    elif "10.1038" in doi:
        url = f"https://www.nature.com/articles/{doi.split('/')[-1]}.pdf"
    elif "10.1126" in doi:
        url = f"https://www.science.org/doi/pdf/{doi}"
    return url, headers

# ---------- NEW DOWNLOAD FUNCTION USING fulltext-article-downloader ----------
def download_one_pdf(doi, scopus_key, wiley_token, download_dir):
    safe_name = doi.replace('/', '_').replace(':', '_') + '.pdf'
    filepath = os.path.join(download_dir, safe_name)
    if os.path.exists(filepath):
        return doi, True, "Already downloaded"

    if not FULLTEXT_AVAILABLE:
        return doi, False, "fulltext-article-downloader not installed"

    try:
        # The library uses environment variables for API keys and email
        # We set them earlier in the main thread; they are inherited by child threads.
        output_path = download_article(doi, output_dir=download_dir, output_filename=safe_name)
        if output_path and os.path.exists(output_path):
            return doi, True, f"Success via fulltext-article-downloader"
        else:
            return doi, False, "Library could not retrieve the article (paywalled or not found)"
    except Exception as e:
        return doi, False, f"Download error: {str(e)}"

# ---------- Keep your existing extraction & AI functions ----------
def extract_pdf_text_fast(pdf_path, max_pages=None, max_chars=20000):
    try:
        doc = fitz.open(pdf_path)
        total_pages = len(doc)
        pages_to_read = doc if max_pages is None or max_pages == 0 else doc[:max_pages]
        text_parts = [page.get_text() for page in pages_to_read if page.get_text()]
        doc.close()
        text = "\n".join(text_parts)
        if max_chars and len(text) > max_chars:
            text = text[:max_chars] + "\n...[truncated]..."
        return text, total_pages
    except:
        return None, 0

def get_cached_text(pdf_path, max_pages, max_chars):
    mtime = os.path.getmtime(pdf_path)
    cache_key = hashlib.md5(f"{pdf_path}_{mtime}_{max_pages}_{max_chars}".encode()).hexdigest()
    cache_file = os.path.join(CACHE_DIR, cache_key)
    if os.path.exists(cache_file):
        try:
            with open(cache_file, "rb") as f:
                return pickle.load(f)
        except:
            pass
    text, pages = extract_pdf_text_fast(pdf_path, max_pages, max_chars)
    if text is not None:
        with open(cache_file, "wb") as f:
            pickle.dump((text, pages), f)
    return text, pages

def get_prompt(prompt_mode, sum_length="standard", template="exec", custom_query=""):
    if prompt_mode == "Quick Summary":
        return f"Summarize in {sum_length} sentences."
    elif prompt_mode == "Template":
        templates = {
            "Executive Summary": "Provide an executive summary (Purpose, Findings, Implications).",
            "Key Findings": "List the key findings with supporting evidence.",
            "Methods": "Describe the methodology and data collection techniques.",
            "Research Questions": "What research questions does this address?",
            "Limitations": "What limitations and future work are mentioned?"
        }
        return templates.get(template, "Summarize.")
    return custom_query if custom_query.strip() else "Summarize."

def call_llm(prompt, backend, model_name=None, ollama_model=None, litellm_model=None, api_key=None, provider=None):
    if backend == "ollama":
        import ollama
        try:
            response = ollama.chat(model=ollama_model, messages=[{"role": "user", "content": prompt}])
            return response["message"]["content"]
        except Exception as e:
            return f"Ollama error: {str(e)}"
    elif backend == "litellm":
        if not LITELLM_AVAILABLE:
            return "LiteLLM not installed. Please run: pip install litellm"
        if not litellm_model or not api_key:
            return "Missing LiteLLM model name or API key."
        if provider == "openai":
            os.environ["OPENAI_API_KEY"] = api_key
        elif provider == "anthropic":
            os.environ["ANTHROPIC_API_KEY"] = api_key
        elif provider == "groq":
            os.environ["GROQ_API_KEY"] = api_key
        elif provider == "google":
            os.environ["GOOGLE_API_KEY"] = api_key
        else:
            os.environ["OPENAI_API_KEY"] = api_key
        try:
            response = completion(model=litellm_model, messages=[{"role": "user", "content": prompt}])
            return response["choices"][0]["message"]["content"]
        except Exception as e:
            return f"LiteLLM error: {str(e)}"
    else:
        return "Unknown backend."

def process_one_pdf_structured(pdf_path, base_prompt, output_spec, backend_config, max_pages, max_chars):
    content, num_pages = get_cached_text(pdf_path, max_pages, max_chars)
    if not content:
        return {"file": os.path.basename(pdf_path), "pages": num_pages, "status": "ERROR", "data": None}
    full_prompt = f"""{base_prompt}

OUTPUT FORMAT INSTRUCTION:
{output_spec}

Based on the above description, generate a JSON object for each paper with the exact column names as keys.
Return a SINGLE JSON object with keys being column names and values being the extracted information.
Example for a specification "Columns: Title, Finding, YES/NO question": 
{{"Title": "...", "Finding": "...", "YES/NO question": "YES"}}
Only return valid JSON. Do not add extra text.

CONTENT:
{content}"""
    raw = call_llm(full_prompt, **backend_config)
    cleaned = re.sub(r'```json\s*|\s*```', '', raw)
    start = cleaned.find('{')
    end = cleaned.rfind('}')
    if start != -1 and end != -1:
        json_str = cleaned[start:end+1]
        try:
            data = json.loads(json_str)
            return {"file": os.path.basename(pdf_path), "pages": num_pages, "status": "OK", "data": data}
        except:
            json_str = re.sub(r',\s*}', '}', json_str)
            json_str = re.sub(r',\s*]', ']', json_str)
            try:
                data = json.loads(json_str)
                return {"file": os.path.basename(pdf_path), "pages": num_pages, "status": "OK", "data": data}
            except:
                pass
    return {"file": os.path.basename(pdf_path), "pages": num_pages, "status": "PARSE_ERROR", "raw": raw, "data": None}

def process_one_pdf_unstructured(pdf_path, base_prompt, backend_config, max_pages, max_chars):
    content, num_pages = get_cached_text(pdf_path, max_pages, max_chars)
    if not content:
        return {"file": os.path.basename(pdf_path), "pages": num_pages, "response": "Extraction failed", "status": "ERROR"}
    full_prompt = f"{base_prompt}\n\nCONTENT:\n{content}"
    resp = call_llm(full_prompt, **backend_config)
    return {"file": os.path.basename(pdf_path), "pages": num_pages, "response": resp, "status": "OK"}

def create_excel_with_formatting(results, output_spec):
    all_keys = set()
    valid_results = [r for r in results if r.get("data")]
    for r in valid_results:
        all_keys.update(r["data"].keys())
    if not all_keys:
        return None
    rows = []
    for r in valid_results:
        row = {"File": r["file"], "Pages": r["pages"]}
        for k in all_keys:
            val = r["data"].get(k, "")
            if isinstance(val, str):
                val_upper = val.strip().upper()
                if val_upper in ["YES", "Y", "TRUE"]:
                    val = "YES"
                elif val_upper in ["NO", "N", "FALSE"]:
                    val = "NO"
            row[k] = val
        rows.append(row)
    df = pd.DataFrame(rows)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Analysis', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Analysis']
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A'
            unique_vals = df[col_name].dropna().astype(str).str.upper().unique()
            if set(unique_vals).issubset({"YES", "NO", ""}):
                start_row = 2
                end_row = start_row + len(df) - 1
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                worksheet.conditional_formatting.add(f"{col_letter}{start_row}:{col_letter}{end_row}",
                                                     FormulaRule(formula=[f'{col_letter}{start_row}="YES"'], fill=green_fill))
                worksheet.conditional_formatting.add(f"{col_letter}{start_row}:{col_letter}{end_row}",
                                                     FormulaRule(formula=[f'{col_letter}{start_row}="NO"'], fill=red_fill))
    output.seek(0)
    return output

def create_csv(results, output_spec):
    all_keys = set()
    valid_results = [r for r in results if r.get("data")]
    for r in valid_results:
        all_keys.update(r["data"].keys())
    if not all_keys:
        return None
    rows = []
    for r in valid_results:
        row = {"File": r["file"], "Pages": r["pages"]}
        for k in all_keys:
            row[k] = r["data"].get(k, "")
        rows.append(row)
    df = pd.DataFrame(rows)
    output = StringIO()
    df.to_csv(output, index=False)
    return output.getvalue()

# ---------- Apply system theme ----------
apply_system_theme()

# ---------- Main UI Layout ----------
st.markdown('<div class="main-header">📚 Researcher OS</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Intelligent academic paper discovery, download, and AI analysis</div>', unsafe_allow_html=True)

# Metric cards
pdf_count = len(st.session_state.pdf_files)
analyzed_count = len(st.session_state.ai_results)
queue_count = len([l for l in st.session_state.dl_queue.splitlines() if l.strip()])

st.markdown('<div class="metric-row">', unsafe_allow_html=True)
col1, col2, col3 = st.columns(3)
with col1:
    st.markdown(f"""
    <div class="metric-card">
        <h3>{pdf_count}</h3>
        <p>📄 PDFs Ready</p>
    </div>
    """, unsafe_allow_html=True)
with col2:
    st.markdown(f"""
    <div class="metric-card">
        <h3>{analyzed_count}</h3>
        <p>✅ Papers Analyzed</p>
    </div>
    """, unsafe_allow_html=True)
with col3:
    st.markdown(f"""
    <div class="metric-card">
        <h3>{queue_count}</h3>
        <p>⚡ Queue</p>
    </div>
    """, unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

# Sidebar
# Sidebar (no theme selector, all expanders start closed)
with st.sidebar:
    st.markdown("## ⚙️ Configuration")
    
    # 1. API Keys (optional) – standalone expander
    with st.expander("🔑 API Keys (optional)", expanded=False):
        scopus_key = st.text_input("Scopus API Key", type="password", key="scopus_key", label_visibility="collapsed", placeholder="Scopus API Key")
        wiley_token = st.text_input("Wiley TDM Token", type="password", key="wiley_token", label_visibility="collapsed", placeholder="Wiley TDM Token")
        unpaywall_email = st.text_input("Unpaywall Email (optional)", key="unpaywall_email", placeholder="your@email.com", label_visibility="collapsed")

        # ---------- Safe environment variable setting ----------
        def clean_env_value(value):
            if not value:
                return None
            cleaned = value.replace('\x00', '').strip()
            return cleaned if cleaned else None

        cleaned_scopus = clean_env_value(scopus_key)
        if cleaned_scopus:
            os.environ["ELSEVIER_API_KEY"] = cleaned_scopus

        cleaned_wiley = clean_env_value(wiley_token)
        if cleaned_wiley:
            os.environ["WILEY_API_KEY"] = cleaned_wiley

        cleaned_email = clean_env_value(unpaywall_email)
        if cleaned_email:
            os.environ["UNPAYWALL_EMAIL"] = cleaned_email

    # 2. LLM Backend – separate expander (not inside API Keys)
    with st.expander("🤖 LLM Backend", expanded=False):
        llm_backend = st.radio("Choose LLM engine:", ["Local Ollama", "Cloud (LiteLLM)"], horizontal=True)
        if llm_backend == "Local Ollama":
            ollama_url = st.text_input("Ollama URL", "http://localhost:11434", label_visibility="collapsed", placeholder="Ollama URL")
            os.environ["OLLAMA_HOST"] = ollama_url
            ollama_model = st.selectbox("Ollama Model", ["llama3.2", "gemma:2b", "phi4", "mistral", "llama3.2:3b"], label_visibility="collapsed")
            backend_config = {"backend": "ollama", "ollama_model": ollama_model}
        else:
            if not LITELLM_AVAILABLE:
                st.error("LiteLLM not installed. Run: pip install litellm")
                st.stop()
            provider = st.selectbox("Provider", ["openai", "anthropic", "groq", "google"], label_visibility="collapsed")
            api_key = st.text_input(f"{provider.upper()} API Key", type="password", label_visibility="collapsed", placeholder="API Key")
            model_name = st.text_input("Model name", placeholder="e.g., gpt-3.5-turbo", label_visibility="collapsed")
            backend_config = {"backend": "litellm", "litellm_model": model_name, "api_key": api_key, "provider": provider}
    
    # 3. Performance – separate expander
    with st.expander("⚡ Performance", expanded=False):
        max_workers_ai = st.slider("Parallel Threads", 1, 8, 4, label_visibility="collapsed")
        max_pages_val = st.number_input("Max Pages (0=all)", 0, 500, 5, label_visibility="collapsed")
        max_chars_val = st.number_input("Max Characters", 1000, 100000, 20000, label_visibility="collapsed")

# Tabs
tab1, tab2, tab3 = st.tabs(["🔍 Discovery", "📥 Download Queue", "🧠 AI Analysis"])

# Tab 1: Discovery
with tab1:
    with st.container():
        col_q, col_l = st.columns([3,1])
        query = col_q.text_input("Search global databases:", placeholder="e.g., 'deep learning' AND 'healthcare'")
        limit = col_l.number_input("Max results (per DB)", 5, 2000, 500)
        if st.button("🚀 Search All", use_container_width=True):
            with st.spinner(f"Fetching up to {limit} papers from OpenAlex..."):
                res = search_openalex(query, limit)
                if scopus_key:
                    with st.spinner(f"Fetching up to {limit} papers from Scopus..."):
                        res += search_scopus(query, scopus_key, limit)
                st.session_state.results_df = pd.DataFrame(res)
                st.success(f"✅ Found {len(res)} papers.")
    
    if st.session_state.results_df is not None:
        st.dataframe(st.session_state.results_df, use_container_width=True)
        if st.button("📌 Add all to Download Queue", use_container_width=True):
            dois = "\n".join(st.session_state.results_df['DOI'].dropna().unique())
            st.session_state.dl_queue += f"{dois}\n"
            st.success("DOIs synced to Tab 2", icon="✅")

# Tab 2: Download Queue
with tab2:
    q_val = st.text_area("Queue (one DOI per line):", value=st.session_state.dl_queue, height=200)
    download_threads = st.slider("Parallel downloads", 1, 10, 5)
    col1, col2 = st.columns(2)
    with col1:
        if st.button("▶️ Start Bulk Download", use_container_width=True):
            if not FULLTEXT_AVAILABLE:
                st.error("fulltext-article-downloader is not installed. Please run: pip install git+https://github.com/computron/fulltext-article-downloader.git")
                st.stop()
            dois = list(set([d.strip() for d in q_val.split('\n') if d.strip()]))
            existing = {f.replace('_', '/').replace('.pdf', '') for f in os.listdir(DOWNLOAD_DIR) if f.endswith('.pdf')}
            dois = [d for d in dois if d not in existing]
            if not dois:
                st.info("All DOIs already downloaded.")
            else:
                progress_bar = st.progress(0)
                status_area = st.empty()
                results = []
                with ThreadPoolExecutor(max_workers=download_threads) as executor:
                    futures = {executor.submit(download_one_pdf, doi, scopus_key, wiley_token, DOWNLOAD_DIR): doi for doi in dois}
                    for i, future in enumerate(as_completed(futures)):
                        doi, success, msg = future.result()
                        results.append(f"{doi}: {msg}")
                        status_area.text(f"Progress: {i+1}/{len(dois)} - Last: {msg}")
                        progress_bar.progress((i+1)/len(dois))
                st.success(f"Download finished. {sum(1 for r in results if 'Success' in r)} succeeded.")
                with st.expander("📜 Download log"):
                    st.write("\n".join(results))
    with col2:
        if st.button("🗑️ Clear Queue", use_container_width=True):
            st.session_state.dl_queue = ""
            st.rerun()

# Tab 3: AI Analysis (unchanged from your original)
with tab3:
    src_folder = st.text_input("📁 Folder to analyze:", value=st.session_state.source_folder)
    if st.button("🔍 Scan for PDFs", use_container_width=True):
        st.session_state.pdf_files = glob.glob(os.path.join(src_folder, "*.pdf"))
        st.session_state.ai_results = []
    st.write(f"📄 PDFs found: **{len(st.session_state.pdf_files)}**")
    
    st.markdown("---")
    st.subheader("✍️ 1. Content Prompt")
    prompt_mode = st.radio("Prompt Mode", ["Quick Summary", "Template", "Custom Query"], horizontal=True)
    col_p1, col_p2 = st.columns(2)
    with col_p1:
        s_len = st.selectbox("Length", ["3", "5", "10"], index=1) if prompt_mode == "Quick Summary" else "standard"
        t_sel = st.selectbox("Template", ["Executive Summary", "Key Findings", "Methods", "Research Questions", "Limitations"]) if prompt_mode == "Template" else "exec"
    with col_p2:
        c_query = st.text_area("Custom Query (free text)") if prompt_mode == "Custom Query" else ""
    base_prompt = get_prompt(prompt_mode, s_len, t_sel, c_query)
    
    st.subheader("📊 2. Output Formatting")
    enable_formatting = st.checkbox("✅ Enable custom output format (table / Excel / PDF)")
    
    if enable_formatting:
        output_spec = st.text_area(
            "Describe the output table layout",
            placeholder="Example: Columns: Paper Title, Main Finding, Supports Hypothesis (YES/NO). Color YES cells green, NO cells red.",
            height=100
        )
        file_type_options = ["Excel (with colors)", "CSV"]
        if REPORTLAB_AVAILABLE:
            file_type_options.append("PDF report")
        file_type = st.selectbox("Export file type", file_type_options)
        
        if st.button("🚀 Start AI Batch Analysis (Formatted)", use_container_width=True):
            if not output_spec.strip():
                st.error("Please provide an output specification.")
            elif not st.session_state.pdf_files:
                st.error("No PDFs found. Click 'Scan for PDFs' first.")
            else:
                if llm_backend == "Cloud (LiteLLM)" and (not backend_config.get("litellm_model") or not backend_config.get("api_key")):
                    st.error("Please provide LiteLLM model name and API key.")
                    st.stop()
                st.session_state.processing = True
                results = []
                p_bar = st.progress(0)
                with ThreadPoolExecutor(max_workers=max_workers_ai) as executor:
                    futures = [executor.submit(process_one_pdf_structured, f, base_prompt, output_spec, backend_config, max_pages_val, max_chars_val) for f in st.session_state.pdf_files]
                    for i, fut in enumerate(as_completed(futures)):
                        results.append(fut.result())
                        p_bar.progress((i+1)/len(futures))
                st.session_state.ai_results = results
                parse_errors = [r for r in results if r["status"] == "PARSE_ERROR"]
                if parse_errors:
                    with st.expander(f"⚠️ {len(parse_errors)} papers had JSON parsing errors"):
                        for err in parse_errors:
                            st.write(f"**{err['file']}**")
                            st.code(err.get('raw', 'No raw output'), language='json')
                if file_type == "Excel (with colors)":
                    excel_data = create_excel_with_formatting(results, output_spec)
                    if excel_data:
                        st.success("Analysis complete. Download the Excel file below.")
                        st.download_button("📥 Download Excel Report", excel_data, "analysis_report.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    else:
                        st.warning("No valid structured data could be extracted.")
                elif file_type == "CSV":
                    csv_data = create_csv(results, output_spec)
                    if csv_data:
                        st.success("Analysis complete. Download the CSV file below.")
                        st.download_button("📥 Download CSV Report", csv_data, "analysis_report.csv", "text/csv")
                    else:
                        st.warning("No valid structured data for CSV export.")
                elif file_type == "PDF report":
                    pdf_data, err = create_pdf_report(results, output_spec)
                    if pdf_data:
                        st.success("Analysis complete. Download the PDF report below.")
                        st.download_button("📥 Download PDF Report", pdf_data, "analysis_report.pdf", "application/pdf")
                    else:
                        st.warning(f"No valid structured data for PDF report. {err if err else ''}")
                st.session_state.processing = False
    else:
        if st.button("🚀 Start AI Batch Analysis (Unstructured)", use_container_width=True):
            if not st.session_state.pdf_files:
                st.error("No PDFs found. Click 'Scan for PDFs' first.")
            else:
                if llm_backend == "Cloud (LiteLLM)" and (not backend_config.get("litellm_model") or not backend_config.get("api_key")):
                    st.error("Please provide LiteLLM model name and API key.")
                    st.stop()
                st.session_state.processing = True
                results = []
                p_bar = st.progress(0)
                with ThreadPoolExecutor(max_workers=max_workers_ai) as executor:
                    futures = [executor.submit(process_one_pdf_unstructured, f, base_prompt, backend_config, max_pages_val, max_chars_val) for f in st.session_state.pdf_files]
                    for i, fut in enumerate(as_completed(futures)):
                        results.append(fut.result())
                        p_bar.progress((i+1)/len(futures))
                st.session_state.ai_results = results
                for res in results:
                    with st.expander(f"📄 {res['file']} ({res.get('pages', '?')} pgs)"):
                        st.write(res['response'])
                if results:
                    df_out = pd.DataFrame(results)
                    st.download_button("💾 Download Analysis CSV", df_out.to_csv(index=False), "analysis.csv", "text/csv")
                st.session_state.processing = False

# Footer
st.markdown("---")
st.caption("Researcher OS v3.5 — System Theme | Powered by OpenAlex, Scopus, Ollama & LiteLLM")